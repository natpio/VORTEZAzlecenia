import streamlit as st
import pandas as pd
from datetime import datetime
import io
import os
import openpyxl
from openpyxl.styles import Alignment
from weasyprint import HTML

# Importujemy silnik Vortex
from core import fetch_data

# --- KONFIGURACJA MAPOWANIA (Zgodnie z załącznikami) ---
MAP_CMR = {
    "1_Nadawca": "D6",               
    "2_Odbiorca": "D14",             
    "3_Miejsce_Przeznaczenia": "D20",
    "4_Miejsce_Zaladunku": "D24",    
    "5_Zalaczone_Dokumenty": "D28",  
    "16_Przewoznik": "M14",          
    "6_Towar": "D32",                
    "11_Waga": "L32",                
    "13_Instrukcje": "D40",          
    "21_Miejsce": "E47",        
    "21_Data": "I47"    
    # O6 usunięte - zachowujemy oryginalny numer seryjny CMR!
}

def fill_excel_and_get_bytes(dane, template_path="cmr_template.xlsx"):
    """Wypełnia fizyczny plik Excel danymi i dba o formatowanie."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    for klucz, komorka in MAP_CMR.items():
        if klucz in dane:
            target_cell = komorka
            # Obsługa scalonych komórek
            for merged_range in ws.merged_cells.ranges:
                if komorka in merged_range:
                    target_cell = str(merged_range).split(':')[0]
                    break
            
            ws[target_cell] = dane[klucz]
            # Wymuszenie zawijania tekstu, żeby długie adresy nie wyjeżdżały poza kratki!
            ws[target_cell].alignment = Alignment(wrapText=True, vertical='top')
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def generate_pro_pdf_cmr(dane):
    """Tworzy profesjonalny PDF (4 strony) na bazie szablonu HTML/CSS."""
    copies = [
        {"color": "#e11d48", "label": "1 - EGZEMPLARZ DLA NADAWCY / SENDER COPY"},
        {"color": "#2563eb", "label": "2 - EGZEMPLARZ DLA ODBIORCY / CONSIGNEE COPY"},
        {"color": "#16a34a", "label": "3 - EGZEMPLARZ DLA PRZEWOŹNIKA / CARRIER COPY"},
        {"color": "#111827", "label": "4 - ADMINISTRACJA / ADMINISTRATION"}
    ]
    
    html_content = ""
    for copy in copies:
        html_content += f"""
        <div class="cmr-page">
            <div class="header">
                <div class="copy-tag" style="background-color: {copy['color']};">{copy['label']}</div>
                <div class="title">INTERNATIONAL CONSIGNMENT NOTE (CMR)</div>
                <div class="subtitle">MIĘDZYNARODOWY LIST PRZEWOZOWY</div>
            </div>
            
            <div class="grid-container">
                <div class="box box-1"><strong>1. Nadawca / Sender:</strong><br>{dane['1_Nadawca'].replace('\n', '<br>')}</div>
                <div class="box box-16"><strong>16. Przewoźnik / Carrier:</strong><br>{dane['16_Przewoznik'].replace('\n', '<br>')}</div>
                <div class="box box-2"><strong>2. Odbiorca / Consignee:</strong><br>{dane['2_Odbiorca'].replace('\n', '<br>')}</div>
                <div class="box box-17"><strong>17. Kolejni przewoźnicy / Successive carriers:</strong><br></div>
                <div class="box box-3"><strong>3. Miejsce rozładunku / Delivery:</strong><br>{dane['3_Miejsce_Przeznaczenia']}</div>
                <div class="box box-4"><strong>4. Miejsce załadunku / Loading:</strong><br>{dane['4_Miejsce_Zaladunku'].replace('\n', '<br>')}</div>
            </div>
            
            <div class="box box-full">
                <strong>6-12. Towar i waga / Description of goods & weight:</strong><br>
                {dane['6_Towar']}<br><br>
                Waga brutto: {dane['11_Waga']}
            </div>
            
            <div class="box box-full"><strong>13. Instrukcje nadawcy / Sender's instructions:</strong><br>{dane['13_Instrukcje']}</div>
            
            <div class="footer-grid">
                <div class="box">21. Wystawiono w:<br>{dane['21_Miejsce']}, {dane['21_Data']}</div>
                <div class="box">22. Podpis nadawcy:</div>
                <div class="box">23. Podpis przewoźnika:</div>
            </div>
            <div class="sys-ref">Zal. dokumenty: {dane['5_Zalaczone_Dokumenty']} | System: Vorteza Orders PRO</div>
        </div>
        """
        
    css = """
    @page { size: A4; margin: 0; }
    body { font-family: Arial, sans-serif; margin: 0; padding: 0; }
    .cmr-page { width: 210mm; height: 297mm; padding: 15mm; border-bottom: 1px dashed #ccc; page-break-after: always; box-sizing: border-box; }
    .header { text-align: right; margin-bottom: 5mm; }
    .copy-tag { color: white; display: inline-block; padding: 3px 10px; font-weight: bold; font-size: 10pt; margin-bottom: 5px; }
    .title { font-size: 16pt; font-weight: bold; color: #333; }
    .subtitle { font-size: 10pt; color: #666; }
    .grid-container { display: table; width: 100%; border-collapse: collapse; }
    .box { border: 1px solid #333; padding: 5px; font-size: 9pt; min-height: 25mm; vertical-align: top; }
    .box-1, .box-2, .box-3, .box-4 { width: 50%; display: table-cell; }
    .box-16, .box-17 { width: 50%; display: table-cell; border-left: none; }
    .box-full { width: 100%; min-height: 40mm; border-top: none; }
    .footer-grid { display: table; width: 100%; border-top: none; }
    .footer-grid .box { display: table-cell; width: 33.3%; height: 30mm; }
    .sys-ref { font-size: 7pt; color: #999; margin-top: 10px; text-align: center; }
    """
    
    full_html = f"<html><head><style>{css}</style></head><body>{html_content}</body></html>"
    pdf_bytes = io.BytesIO()
    HTML(string=full_html).write_pdf(pdf_bytes)
    pdf_bytes.seek(0)
    return pdf_bytes

# --- INTERFEJS TERMINALA ---
st.markdown("<h2 style='color: #38bdf8;'>📄 TERMINAL CMR (AUTO-EXCEL)</h2>", unsafe_allow_html=True)

if not os.path.exists("cmr_template.xlsx"):
    st.error("Wgraj plik cmr_template.xlsx do głównego folderu aplikacji na GitHubie!")
    st.stop()

with st.spinner("Ładowanie zleceń..."):
    df = fetch_data("Zlecenia")

if not df.empty:
    if 'Dział' in df.columns:
        df_cargo = df[df['Dział'] == 'LOGISTYKA CARGO']
    elif len(df.columns) > 2:
        col_dzial = df.columns[2]
        df_cargo = df[df[col_dzial] == 'LOGISTYKA CARGO']
    else:
        df_cargo = df
        
    if not df_cargo.empty and 'Numer zlecenia' in df_cargo.columns:
        lista_nr = df_cargo['Numer zlecenia'].astype(str).tolist()
    else:
        lista_nr = []

    if lista_nr:
        wybor = st.selectbox("Wybierz zlecenie do CMR:", ["Wybierz..."] + lista_nr)

        if wybor != "Wybierz...":
            r = df_cargo[df_cargo['Numer zlecenia'].astype(str) == wybor].iloc[0]
            
            uwagi_col = 'Uwagi / Instrukcje' if 'Uwagi / Instrukcje' in df.columns else (df.columns[13] if len(df.columns) > 13 else 'Brak')
            uwagi_raw = str(r.get(uwagi_col, ''))
            pojazd = uwagi_raw.split('AUTO: ')[-1].split(' ||')[0] if 'AUTO: ' in uwagi_raw else "TBA"
            
            # Waga z inputa, żeby można było poprawić ręcznie
            waga_input = st.text_input("Waga brutto do CMR:", value="Zgodnie ze specyfikacją")
            
            # Przygotowanie danych z uwzględnieniem danych z Twojego pliku
            dane_doc = {
                "1_Nadawca": "SQM Prosta Spółka Akcyjna\nul. Poznańska 165, 62-052 Komorniki\nNIP: 7792361182",
                "2_Odbiorca": f"TARGI / EVENT: {r.get('ID Projektu', 'N/A')}\n{r.get('Miejsce Rozladunku', '')}",
                "3_Miejsce_Przeznaczenia": str(r.get('Miejsce Rozladunku', '')),
                "4_Miejsce_Zaladunku": f"{r.get('Miejsce Zaladunku', '')}\nData: {r.get('Data Zaladunku', '')}",
                "5_Zalaczone_Dokumenty": f"Zlecenie: {wybor}",
                "16_Przewoznik": f"{r.get('Zleceniobiorca', '')}\nAuto: {pojazd}",
                "6_Towar": "Exhibition Structures / Sprzęt AV",
                "11_Waga": waga_input,
                "13_Instrukcje": uwagi_raw,
                "21_Miejsce": "Komorniki",
                "21_Data": datetime.now().strftime('%Y-%m-%d')
            }

            col_ex, col_pdf = st.columns(2)
            
            with col_ex:
                if st.button("📊 PRZYGOTUJ EXCEL (.xlsx)", use_container_width=True):
                    with st.spinner("Wypełnianie szablonu..."):
                        try:
                            ex_file = fill_excel_and_get_bytes(dane_doc)
                            st.download_button("📥 POBIERZ WYPEŁNIONY EXCEL", data=ex_file, file_name=f"CMR_{wybor.replace('/', '_')}.xlsx")
                        except Exception as e:
                            st.error(f"Błąd excela: {e}")

            with col_pdf:
                if st.button("⚡ GENERUJ PDF PRO", type="primary", use_container_width=True):
                    with st.spinner("Przetwarzanie dokumentu (WeasyPrint)..."):
                        try:
                            pdf_file = generate_pro_pdf_cmr(dane_doc)
                            st.download_button("📥 POBIERZ CMR PDF (4 STRONY)", data=pdf_file, file_name=f"CMR_{wybor.replace('/', '_')}.pdf")
                        except Exception as e:
                            st.error(f"Błąd PDF: {e}")
    else:
        st.warning("Nie znaleziono zleceń w bazie dla podanego działu.")
else:
    st.error("Baza zleceń jest pusta.")
